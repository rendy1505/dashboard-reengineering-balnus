#%%
import openpyxl

DEP_PATH = "/Users/mac/TELKOMSEL - TI BALNUSRA/Dashboard Reengineering Balnus/Source Data Dashboard/Deployment/DeploymentManagement_EQUIPMENT_20260805_140510_2025_2026.xlsx"
MAP_PATH = "/Users/mac/TELKOMSEL - TI BALNUSRA/Dashboard Reengineering Balnus/Source Data Dashboard/Sitelist Donor - Acceptor/Catatan program TI mery.v1.xlsx"

#%% Build Project SOW ID -> EQP Donor lookup from the "EQP Donor" sheet
wb_map = openpyxl.load_workbook(MAP_PATH, read_only=True, data_only=True)
ws_map = wb_map["EQP Donor"]

header_map = [c.value for c in next(ws_map.iter_rows(min_row=1, max_row=1))]
col_map = {name: i for i, name in enumerate(header_map) if name}

sow_i = col_map["Project SOW ID"]
eqp_i = col_map["EQP Donor"]

# Last row wins if a SOW ID repeats with a different EQP Donor value.
eqp_donor_lookup = {}
for row in ws_map.iter_rows(min_row=2, values_only=True):
    sow_id = row[sow_i]
    eqp_donor = row[eqp_i]
    if not sow_id or not eqp_donor:
        continue
    eqp_donor_lookup[str(sow_id).strip()] = str(eqp_donor).strip()

wb_map.close()
print(f"EQP Donor: {len(eqp_donor_lookup)} unique Project SOW ID entries loaded")

#%% Load Deployment workbook (keep formatting/formulas — not read_only) and
# add the "EQP Donor" column.
wb_dep = openpyxl.load_workbook(DEP_PATH)
ws_dep = wb_dep["Sheet1"]

header_dep = [c.value for c in next(ws_dep.iter_rows(min_row=1, max_row=1))]
col_dep = {name: i + 1 for i, name in enumerate(header_dep)}  # 1-indexed for openpyxl cell access

if "EQP Donor" not in col_dep:
    col_dep["EQP Donor"] = ws_dep.max_column + 1
    ws_dep.cell(row=1, column=col_dep["EQP Donor"], value="EQP Donor")

eqp_donor_col = col_dep["EQP Donor"]
sow_col = col_dep["Project SOW ID"]

matched = 0
unmatched = 0

for r in range(2, ws_dep.max_row + 1):
    sow_id = ws_dep.cell(row=r, column=sow_col).value
    hit = eqp_donor_lookup.get(str(sow_id).strip()) if sow_id else None
    if hit:
        ws_dep.cell(row=r, column=eqp_donor_col, value=hit)
        matched += 1
    else:
        ws_dep.cell(row=r, column=eqp_donor_col, value=None)
        unmatched += 1

print(f"Matched: {matched}  |  Unmatched (left blank): {unmatched}")

wb_dep.save(DEP_PATH)
wb_dep.close()
print("Saved:", DEP_PATH)

#%% Sample: show 1 matched site's row to sanity-check
wb_check = openpyxl.load_workbook(DEP_PATH, read_only=True, data_only=True)
ws_check = wb_check["Sheet1"]
header_check = [c.value for c in next(ws_check.iter_rows(min_row=1, max_row=1))]
col_check = {name: i for i, name in enumerate(header_check)}

sample_row = None
for row in ws_check.iter_rows(min_row=2, values_only=True):
    if row[col_check["EQP Donor"]]:
        sample_row = row
        break

cols_to_show = ["Project SOW ID", "Site Id", "EQP Donor"]
print(f"{'Column':<18} | Value")
print("-" * 50)
for name in cols_to_show:
    print(f"{name:<18} | {sample_row[col_check[name]]}")
wb_check.close()
