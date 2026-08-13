#%%
import openpyxl

DEP_PATH = "/Users/mac/TELKOMSEL - TI BALNUSRA/Dashboard Reengineering Balnus/Source Data Dashboard/Deployment/DeploymentManagement_EQUIPMENT_20260805_140510_2025_2026.xlsx"
MAP_PATH = "/Users/mac/TELKOMSEL - TI BALNUSRA/Dashboard Reengineering Balnus/Source Data Dashboard/Sitelist Mapping/BOQ New Presales_mapping check.xlsx"

NEED_UPDATE = "Need Update"

#%% Build SOW ID -> (Site ID Acceptor, Remark) lookup from the mapping file
# "mapping Ok" has two separate blocks of SOW ID/NEID/SECTOR columns; the
# acceptor-mapping block (SOW ID, SITE_ID New Acceptor, NEID, SECTOR, Note,
# Remark) is the SECOND occurrence of each name in the header row, so
# indexing columns by name (later occurrence overwrites earlier one in the
# dict) naturally lands on that block instead of the first (SITE_ID /
# SITE_ID DONOR) block.
wb_map = openpyxl.load_workbook(MAP_PATH, read_only=True, data_only=True)
ws_map = wb_map["mapping Ok"]

header_map = [c.value for c in next(ws_map.iter_rows(min_row=1, max_row=1))]
col_map = {name: i for i, name in enumerate(header_map) if name}

sow_i = col_map["SOW ID"]
acc_i = col_map["SITE_ID New Acceptor"]
remark_i = col_map["Remark"]

# Last row wins if a SOW ID repeats.
acceptor_lookup = {}
for row in ws_map.iter_rows(min_row=2, values_only=True):
    sow_id = row[sow_i]
    if not sow_id:
        continue
    acceptor = row[acc_i]
    remark = row[remark_i]
    acceptor_lookup[str(sow_id).strip()] = (
        str(acceptor).strip() if acceptor else None,
        str(remark).strip() if remark else None,
    )

wb_map.close()
print(f"mapping Ok: {len(acceptor_lookup)} unique SOW ID entries loaded")

#%% Load Deployment workbook (keep formatting/formulas — not read_only) and
# add the 2 new columns: Remark Mapping, Site ID Acceptor Mapping (both from
# the mapping file above). Site ID Donor is deliberately NOT added yet —
# still being discussed — see chat.
wb_dep = openpyxl.load_workbook(DEP_PATH)
ws_dep = wb_dep["Sheet1"]

header_dep = [c.value for c in next(ws_dep.iter_rows(min_row=1, max_row=1))]
col_dep = {name: i + 1 for i, name in enumerate(header_dep)}  # 1-indexed for openpyxl cell access


def ensure_col(name):
    if name not in col_dep:
        col_dep[name] = ws_dep.max_column + 1
        ws_dep.cell(row=1, column=col_dep[name], value=name)
    return col_dep[name]


remark_col = ensure_col("Remark Mapping")
acceptor_col = ensure_col("Site ID Acceptor Mapping")

sow_col = col_dep["Project SOW ID"]

matched = 0
unmatched = 0

for r in range(2, ws_dep.max_row + 1):
    sow_id = ws_dep.cell(row=r, column=sow_col).value

    hit = acceptor_lookup.get(str(sow_id).strip()) if sow_id else None
    if hit:
        acceptor, remark = hit
        ws_dep.cell(row=r, column=acceptor_col, value=acceptor)
        ws_dep.cell(row=r, column=remark_col, value=remark)
        matched += 1
    else:
        ws_dep.cell(row=r, column=acceptor_col, value=NEED_UPDATE)
        ws_dep.cell(row=r, column=remark_col, value=NEED_UPDATE)
        unmatched += 1

print(f"Matched: {matched}  |  Unmatched (marked '{NEED_UPDATE}'): {unmatched}")

wb_dep.save(DEP_PATH)
wb_dep.close()
print("Saved:", DEP_PATH)

#%% Sample: show 1 matched site's row to sanity-check
wb_check = openpyxl.load_workbook(DEP_PATH, read_only=True, data_only=True)
ws_check = wb_check["Sheet1"]
header_check = [c.value for c in next(ws_check.iter_rows(min_row=1, max_row=1))]
col_check = {name: i for i, name in enumerate(header_check)}

sample_row = None
for row in ws_check.iter_rows(min_row=2, values_only=True):
    if row[col_check["Site ID Acceptor Mapping"]] not in (None, NEED_UPDATE):
        sample_row = row
        break

cols_to_show = ["Project SOW ID", "Site Id", "Site ID Acceptor Mapping", "Remark Mapping"]
print(f"{'Column':<28} | Value")
print("-" * 60)
for name in cols_to_show:
    print(f"{name:<28} | {sample_row[col_check[name]]}")
wb_check.close()
